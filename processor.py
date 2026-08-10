import streamlit as st
import openpyxl
import pdfplumber
import re
import requests
import json
from io import BytesIO

from parser_bkt_register import extract_bkt_items, map_items_to_excel_dynamic as map_bkt
from shipper_data import ensure_default_shipper
from pdf_engine import extract_header_value

GOOGLE_SHEET_WEB_APP_URL = "https://script.google.com/macros/s/AKfycbwYVVWbqNZbzTOujVmip41KlID-rf9zEQLy_JM04ZEhUL-kixwRMD9nbPnOrZ46Fmz3/exec"

def send_data_to_target_google_sheet(sheet_link, tab_name, wb):
    """यदि शिपर के पास टारगेट गूगल शीट लिंक है तो एक्सेल डेटा को शीट पर पुश करता है"""
    try:
        if not sheet_link or not sheet_link.strip():
            return False
            
        ws = wb["INV"] if "INV" in wb.sheetnames else wb.active
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            if any(row): 
                rows_data.append(list(row))
                
        payload = {
            "action": "append_to_target_sheet",
            "sheet_link": sheet_link.strip(),
            "tab_name": tab_name.strip() if tab_name else "Sheet1",
            "data": rows_data
        }
        
        response = requests.post(GOOGLE_SHEET_WEB_APP_URL, data=json.dumps(payload), timeout=60)
        return response.status_code == 200
    except Exception:
        return False

def render_processor():
    ensure_default_shipper()
    
    with st.sidebar:
        st.markdown("---")
        try:
            st.image("ck_photo.jpg", width=100, caption="Chetan Joshi")
        except Exception:
            st.markdown("👤")
            
        st.markdown("### **System Developer**")
        st.markdown("**Chetan Joshi**")
        st.markdown("📞 +91 98253 06898")
        st.markdown("---")
    
    st.header("📄 CK PDF PROCESSOR")
    st.caption("एक साथ कई स्टैंडर्ड PDF इनवॉइस अपलोड करें, प्रोसेस करें और एक्सेल डाउनलोड करें।")
    
    shippers_list = sorted(list(st.session_state["shipper_database"].keys()))
    
    selected_shipper = st.selectbox(
        "किस शिपर का इनवॉइस प्रोसेस करना है?", 
        shippers_list,
        index=None,
        placeholder="शिपर का नाम टाइप करें या चुनें...",
        key="processor_shipper_select"
    )
    
    if not selected_shipper:
        st.info("👆 कृपया प्रोसेसिंग शुरू करने के लिए ऊपर दिए गए ड्रॉपडाउन से शिपर चुनें।")
        return
        
    if selected_shipper:
        shipper_info = st.session_state["shipper_database"][selected_shipper]
        assigned_parser = shipper_info.get("item_table_rule_name", "parser_bkt_register")
        
        st.markdown("---")
        st.subheader("📂 Custom Excel Format / Template (Optional)")
        excel_template = st.file_uploader(
            "यदि आपके पास कोई अपना फिक्स एक्सेल फॉर्मेट है, तो यहाँ अपलोड करें:", 
            type=["xlsx"], 
            key="processor_excel_template"
        )
        
        st.markdown("---")
        st.subheader("📑 Upload PDF's")
        
        uploaded_pdfs = st.file_uploader(
            "एक या एक से अधिक स्टैंडर्ड इनवॉइस PDF चुनें", 
            type=["pdf"], 
            accept_multiple_files=True,
            key=f"multi_pdf_uploader_{selected_shipper}"
        )
        
        if uploaded_pdfs:
            st.success(f"🎉 कुल {len(uploaded_pdfs)} PDF फाइलें सफलतापूर्वक सेलेक्ट हो गई हैं!")
            
            st.write("---")
            st.markdown("#### ⚡ Processing Actions")
            
            if st.button("🚀 Process All PDFs", type="primary", use_container_width=True):
                with st.spinner(f"कुल {len(uploaded_pdfs)} PDF फाइलें प्रोसेस हो रही हैं... कृपया प्रतीक्षा करें..."):
                    rules = shipper_info.get("mapping_rules", {})
                    item_table_rules = shipper_info.get("item_table_rules", {})
                    
                    target_sheet_link = shipper_info.get("target_sheet_link", "")
                    target_tab_name = shipper_info.get("target_tab_name", "Sheet1")
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "INV"
                    
                    if excel_template:
                        wb = openpyxl.load_workbook(excel_template)
                        ws = wb["INV"] if "INV" in wb.sheetnames else wb.active
                        
                        # सटीक अंतिम भरी हुई रो ढूँढना (जैसे स्क्रीनशॉट में 14)
                        last_filled_row = 0
                        for r in range(ws.max_row, 0, -1):
                            if any(ws.cell(row=r, column=c).value is not None for c in range(1, ws.max_column + 1)):
                                last_filled_row = r
                                break
                        excel_write_row = last_filled_row + 1 if last_filled_row > 0 else 2
                    else:
                        # बिना टेम्पलेट के: Row 1 में हेडिंग्स लिखें
                        excel_write_row = 2
                        col_idx = 1
                        for field_name in rules.keys():
                            ws.cell(row=1, column=col_idx, value=field_name)
                            col_idx += 1
                        # डिफ़ॉल्ट सिस्टम कॉलम हेडिंग्स
                        ws.cell(row=1, column=col_idx, value="SR NO")
                        ws.cell(row=1, column=col_idx+1, value="INVOICE NO")
                        ws.cell(row=1, column=col_idx+2, value="INVOICE DATE")
                    
                    first_inv_no = "INV"
                    overall_item_sr = 1
                    current_excel_row = excel_write_row
                    
                    for inv_idx, inv_file in enumerate(uploaded_pdfs):
                        pdf_text = ""
                        pdf_lines = []
                        
                        file_bytes_cache = inv_file.getvalue()
                        st.session_state["cached_pdf_bytes"] = file_bytes_cache
                        
                        with pdfplumber.open(BytesIO(file_bytes_cache)) as pdf:
                            for page in pdf.pages:
                                t = page.extract_text()
                                if t:
                                    pdf_text += t + "\n"
                                    pdf_lines.extend(t.split("\n"))
                        
                        current_inv_number = f"INV_{inv_idx+1}"
                        current_inv_date = ""
                        summary_row = current_excel_row
                        
                        # हेडर फील्ड्स मैपिंग और वैल्यू एक्सट्रैक्शन
                        field_col_counter = 1
                        for field, r_info in rules.items():
                            kw = r_info.get("keyword", "").strip()
                            if kw.startswith("'") and len(kw) > 1:
                                kw = kw[1:].strip()
                                
                            pos = r_info.get("position", "Right (आगे)")
                            target_cell = r_info.get("cell", "").strip().upper()
                            mode = r_info.get("match_mode", "Exact Word")
                            stop_kw = r_info.get("stop_kw", "")
                            flt = r_info.get("filter", "None")
                            fallback_val = r_info.get("fallback", "").strip()
                            
                            found_val = extract_header_value(pdf_lines, pdf_text, kw, pos, mode, stop_kw, flt, field_label=field, pdf_bytes=file_bytes_cache)
                            
                            if not found_val or not found_val.strip():
                                if fallback_val:
                                    found_val = fallback_val
                                    
                            # यदि यूजर ने सेल लेटर (जैसे C, L) दिया है तो वहाँ लिखें, अन्यथा बिना टेम्पलेट के सीक्वेसल कॉलम में लिखें
                            if target_cell and "dynamic" not in target_cell.lower():
                                try:
                                    if target_cell.isalpha():
                                        cell_to_write = f"{target_cell}{summary_row}"
                                    else:
                                        cell_to_write = target_cell
                                    ws[cell_to_write] = found_val
                                except Exception:
                                    pass
                            elif not excel_template:
                                ws.cell(row=summary_row, column=field_col_counter, value=found_val)
                            
                            field_col_counter += 1
                            
                            if "inv. no" in field.lower() or "invoice no" in field.lower():
                                if found_val:
                                    current_inv_number = found_val
                                    if inv_idx == 0: first_inv_no = found_val
                            
                            if "date" in field.lower() or "dt" in field.lower():
                                d_match = re.search(r'\b\d{2}[./-]\d{2}[./-]\d{4}\b', found_val)
                                if d_match:
                                    current_inv_date = d_match.group(0).replace(".", "/").replace("-", "/")
                                elif found_val and not found_val.lower().startswith("inv"):
                                    current_inv_date = found_val

                        # डिफ़ॉल्ट ट्रेलिंग डेटा (SR NO, Invoice No, Date)
                        if not excel_template:
                            ws.cell(row=summary_row, column=field_col_counter, value=inv_idx + 1)
                            ws.cell(row=summary_row, column=field_col_counter+1, value=current_inv_number)
                            if current_inv_date:
                                ws.cell(row=summary_row, column=field_col_counter+2, value=current_inv_date)
                        else:
                            ws[f"AH{summary_row}"] = inv_idx + 1
                            ws[f"AI{summary_row}"] = current_inv_number
                            if current_inv_date:
                                ws[f"AJ{summary_row}"] = current_inv_date

                        resolved_item_rules = {}
                        for i_name, i_info in item_table_rules.items():
                            i_type = i_info.get("type", "")
                            i_rule = i_info.get("rule", "")
                            i_col = i_info.get("col", "K")
                            resolved_item_rules[i_name] = {"col": i_col, "type": i_type, "rule": i_rule}

                        # पार्सर के माध्यम से आइटम्स प्रोसेस करना और अगली रो अपडेट करना
                        parsed_items = extract_bkt_items(pdf_lines, pdf_text=pdf_text)
                        ws, overall_item_sr, current_excel_row = map_bkt(
                            ws, parsed_items, resolved_item_rules,
                            inv_sr_no=inv_idx+1, 
                            start_overall_sr=overall_item_sr, 
                            start_excel_row=summary_row, 
                            default_invoice_no=current_inv_number, 
                            default_invoice_date=current_inv_date,
                            pdf_text=pdf_text,
                            parser_rule=assigned_parser
                        )

                    output = BytesIO()
                    wb.save(output)
                    
                    if target_sheet_link and target_sheet_link.strip():
                        sheet_success = send_data_to_target_google_sheet(target_sheet_link, target_tab_name, wb)
                        if sheet_success:
                            st.success("☁️ डेटा आपकी दी गई टारगेट गूगल शीट पर भी सफलतापूर्वक सिंक हो गया है!")
                        else:
                            st.warning("⚠️ एक्सेल फाइल तैयार है, लेकिन टारगेट गूगल शीट पर डेटा भेजने में विफल रहा (लिंक चेक करें)।")

                    short_shipper = selected_shipper.split(" ")[0].lower()
                    clean_inv = re.sub(r'[\\/*?:"<>|]', "", first_inv_no)
                    final_filename = f"{clean_inv}_{short_shipper}_BatchProcessed.xlsx"
                    
                    st.session_state["processed_file_ready"] = {"filename": final_filename, "data": output.getvalue()}
                    st.success("🎉 सभी PDF सफलतापूर्वक प्रोसेस हो गए हैं! अब नीचे डाउनलोड बटन से फाइल डाउनलोड करें।")
            
            if st.session_state.get("processed_file_ready", None):
                st.write("---")
                st.download_button(
                    label=f"📥 Download Excel ({st.session_state['processed_file_ready']['filename']})",
                    data=st.session_state['processed_file_ready']['data'],
                    file_name=st.session_state['processed_file_ready']['filename'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
